from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import json

def flatten_colegio_data(colegio):
    """Aplana la estructura anidada del colegio en un diccionario plano"""
    flat_data = {}
    
    # Datos generales
    for key, value in colegio['general'].items():
        flat_data[f'general_{key}'] = value
    
    # Ubicación
    for key, value in colegio['ubicacion'].items():
        if key == 'coordenadas':
            flat_data['latitud'] = value['latitud']
            flat_data['longitud'] = value['longitud']
            flat_data['coordenadas_texto'] = value['texto']
        else:
            flat_data[f'ubicacion_{key}'] = value
    
    # Estadísticas
    stats = colegio['estadisticas']
    for stat_type, categories in stats.items():
        for category, years in categories.items():
            for year, value in years.items():
                col_name = f"estadistica_{stat_type}_{category}_{year}"
                flat_data[col_name] = value
    
    # Infraestructura
    infra = colegio['infraestructura']
    for section, items in infra.items():
        for item, value in items.items():
            col_name = f"infra_{section}_{item}"
            flat_data[col_name] = value
    
    flat_data['url_fuente'] = colegio['url']
    return flat_data

def export_to_excel(colegios_data, filename='colegios_data.xlsx'):
    """Exporta los datos a Excel usando solo openpyxl"""
    if not colegios_data:
        print("No hay datos para exportar")
        return
    
    # Aplanar todos los datos
    flattened_data = [flatten_colegio_data(colegio) for colegio in colegios_data]
    
    # Obtener todas las claves únicas (columnas)
    all_keys = set()
    for data in flattened_data:
        all_keys.update(data.keys())
    columns = sorted(all_keys)
    
    # Crear libro de Excel
    wb = Workbook()
    ws = wb.active
    
    # Escribir encabezados
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col_num, column_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num, value=column_name)
        cell.fill = header_fill
        cell.font = header_font
    
    # Escribir datos
    for row_num, colegio_data in enumerate(flattened_data, 2):
        for col_num, column_name in enumerate(columns, 1):
            value = colegio_data.get(column_name, "")
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Autoajustar columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                cell_value = str(cell.value) if cell.value is not None else ""
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Guardar archivo
    wb.save(filename)
    print(f"\nArchivo Excel generado exitosamente: {filename}")
    print(f"Total de colegios exportados: {len(colegios_data)}")
    print(f"Total de columnas: {len(columns)}")

# Ejemplo de uso (debes integrarlo con tu código principal)
if __name__ == "__main__":
    # Aquí cargarías tus datos reales
    with open('colegios_data.json', 'r', encoding='utf-8') as f:
        colegios_data = json.load(f)
    
    export_to_excel(colegios_data)