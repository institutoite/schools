import re
import time
import requests
from urllib.parse import urljoin
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# ================== CONFIG ==================
BASE = "https://cm.organojudicial.gob.bo/consejo/manualprocedimientosddrr/"
START_URL = urljoin(BASE, "index.html")
OUT_XLSX = "reporte_ortografia_manual.xlsx"

MAX_PAGES = 800
SLEEP_PAGE = 0.5
TIMEOUT = 25

# LanguageTool (API pública)
LT_API_URL = "https://api.languagetool.org/v2/check"
LANGUAGE = "es"
CHUNK_SIZE = 8000

session = requests.Session()
session.headers.update({
    "User-Agent": "ManualSpellChecker/1.0 (uso editorial)"
})

# ================== HELPERS ==================
def normalize_url(u: str) -> str:
    return u.split("#")[0]

def is_in_manual(u: str) -> bool:
    return u.startswith(BASE)

def fetch_soup(url: str) -> BeautifulSoup:
    r = session.get(url, timeout=TIMEOUT)
    r.raise_for_status()
    return BeautifulSoup(r.text, "lxml")

def extract_sidebar_links(soup: BeautifulSoup, current_url: str) -> set[str]:
    links = set()
    sidebar_candidates = soup.select("nav, aside, .sidebar, .sidenav, .toc, .navigation, #sidebar, #nav")
    containers = sidebar_candidates if sidebar_candidates else [soup]

    for c in containers:
        for a in c.select("a[href]"):
            href = (a.get("href") or "").strip()
            if not href or href.startswith("#"):
                continue
            full = normalize_url(urljoin(current_url, href))
            if is_in_manual(full):
                links.add(full)
    return links

def extract_main_text(soup: BeautifulSoup) -> str:
    main = soup.select_one("main, article, .content, .document, .page, .container") or (soup.body or soup)
    text = main.get_text("\n", strip=True)

    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{2,}", "\n", text)
    return text

def lt_check_text(text: str) -> list[dict]:
    matches_all = []
    offset_base = 0

    for i in range(0, len(text), CHUNK_SIZE):
        chunk = text[i:i + CHUNK_SIZE]
        resp = session.post(
            LT_API_URL,
            data={"text": chunk, "language": LANGUAGE},
            timeout=TIMEOUT
        )
        resp.raise_for_status()
        data = resp.json()

        for m in data.get("matches", []):
            m2 = dict(m)
            m2["offset"] = int(m2.get("offset", 0)) + offset_base
            matches_all.append(m2)

        offset_base += len(chunk)
        time.sleep(0.15)

    return matches_all

def snippet(text: str, offset: int, length: int, window: int = 70) -> str:
    a = max(0, offset - window)
    b = min(len(text), offset + length + window)
    return text[a:b].replace("\n", " ")

# ================== CRAWL ==================
def crawl_all_pages_from_sidebar() -> set[str]:
    discovered = set()
    queue = [START_URL]

    while queue and len(discovered) < MAX_PAGES:
        url = queue.pop(0)
        if url in discovered:
            continue

        try:
            soup = fetch_soup(url)
        except Exception:
            discovered.add(url)
            continue

        discovered.add(url)

        new_links = extract_sidebar_links(soup, url)
        for lk in new_links:
            if lk not in discovered and lk not in queue:
                queue.append(lk)

        time.sleep(SLEEP_PAGE)

    return discovered

# ================== EXCEL ==================
def autosize_columns(ws):
    for col in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for cell in ws[col_letter]:
            val = "" if cell.value is None else str(cell.value)
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 80)

def run_audit_to_excel(pages: set[str]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Errores"

    # Columnas pedidas: URL | Error | Correcto (sugerencia)
    # Recomendado: Contexto para ubicar
    ws.append(["Pagina (URL)", "Error detectado", "Sugerencia (correcto)", "Contexto"])

    for url in sorted(pages):
        try:
            soup = fetch_soup(url)
            text = extract_main_text(soup)

            if len(text) < 80:
                continue

            matches = lt_check_text(text)

            for m in matches:
                off = int(m.get("offset", 0))
                length = int(m.get("length", 0))
                message = m.get("message", "")

                repls = m.get("replacements", [])
                suggestion = repls[0].get("value", "") if repls else ""

                ctx = snippet(text, off, length)

                ws.append([url, message, suggestion, ctx])

            time.sleep(SLEEP_PAGE)

        except Exception as e:
            ws.append([url, f"ERROR: {e}", "", ""])

    autosize_columns(ws)
    wb.save(OUT_XLSX)
    print(f"Listo: {OUT_XLSX}")

if __name__ == "__main__":
    pages = crawl_all_pages_from_sidebar()
    print(f"Encontradas {len(pages)} páginas.")
    run_audit_to_excel(pages)